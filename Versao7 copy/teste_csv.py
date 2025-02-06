#from extractors_utils import extract_text_from_csv, extract_text_and_images_from_docx
from io import BytesIO
import os
import pandas as pd

# file_path = r"D:\Repositorios\ocr-api\Versao7\arquivosGerais\docx\OP_90_2024_ANOT_AN_TELEMÁTICOS4.docx"

# texto, sucesso, erro_msg = extract_text_from_csv(file_path)

# if sucesso:
#     with open("output.txt", "w", encoding="utf-8") as file:
#         file.write(texto)
#     print("Texto extraído com sucesso:")
#     #print(texto)
# else:
#     print(f"Erro durante a extração: {erro_msg}")



# ### Tratamento XLS

import pandas as pd
from openpyxl import load_workbook
import xlrd
import mimetypes
import os

file_path = r"D:\Repositorios\ocr-api\Versao7\arquivosGerais\csvs\OP_213_2023_OP_SENHA_VG.xlsx"

# def process_excel_file(file_path):
#     """
#     Lê todas as abas de um arquivo Excel (.xls ou .xlsx) e retorna o texto extraído.

#     Args:
#         file_path (str): Caminho para o arquivo Excel.

#     Returns:
#         str: Texto consolidado de todas as abas do arquivo ou mensagem de erro.
#     """
#     try:
#         # Verifica se o arquivo existe
#         if not os.path.exists(file_path):
#             return f"Erro: O arquivo {file_path} não foi encontrado."

#         # Detecta o tipo do arquivo
#         mime_type, _ = mimetypes.guess_type(file_path)
#         texto = ""

#         if mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
#             # Trata arquivos .xlsx
#             print("Detectado formato XLSX.")
#             workbook = load_workbook(filename=file_path, data_only=True)
#             for sheet_name in workbook.sheetnames:
#                 texto += f"\n=== Aba: {sheet_name} ===\n"
#                 sheet = workbook[sheet_name]
#                 for row in sheet.iter_rows(values_only=True):
#                     if any(row):  # Garante que pelo menos uma célula tem valor
#                         texto += ", ".join(map(str, row)) + "\n"
#             return texto.strip()

#         elif mime_type == 'application/vnd.ms-excel':
#             # Trata arquivos .xls
#             print("Detectado formato XLS.")
#             workbook = xlrd.open_workbook(file_path)
#             for sheet in workbook.sheets():
#                 texto += f"\n=== Aba: {sheet.name} ===\n"
#                 for row_idx in range(sheet.nrows):
#                     row = sheet.row_values(row_idx)
#                     if any(row):  # Garante que pelo menos uma célula tem valor
#                         texto += ", ".join(map(str, row)) + "\n"
#             return texto.strip()

#         else:
#             return f"Erro: Formato do arquivo não reconhecido ({mime_type})."

#     except xlrd.biffh.XLRDError as e:
#         return f"Erro ao processar arquivo XLS: {e}"
#     except Exception as e:
#         return f"Erro ao processar arquivo: {e}"


# def extract_text_from_xls(file_path):
#     """
#     Função mestre que tenta ler um arquivo como planilha ou similar.

#     Args:
#         file_path (str): Caminho para o arquivo XLS ou similar.

#     Returns:
#         str: Texto extraído do arquivo.
#     """
#     try:
#         return process_excel_file(file_path)
#     except Exception as e:
#         return f"Erro ao processar arquivo: {e}"


# # Testando o script
# resultado = extract_text_from_xls(file_path)
# with open(r"D:\Repositorios\ocr-api\Versao7\arquivosGerais\csvs\output.txt", "w", encoding="utf-8") as file:
#     file.write(resultado)
# #print(resultado)


# ### Fim Tratamento XLS


# Caminho para a imagem
#image_path = r"D:\Repositorios\ocr-api\Versao7\images\24_34_17022020084236-285.jpg"

# # Extrai o texto da imagem
# texto, sucesso, erro_msg = extract_text_from_image(image_path)

# if sucesso:
#     print("Texto extraído com sucesso:")
#     print(texto)
# else:
#     print(f"Erro durante a extração: {erro_msg}")


# Caminho para o arquivo DOCX
# docx_path = r"D:\Repositorios\ocr-api\Versao7\arquivosGerais\docx\OP_90_2024_ANOT_AN_TELEMÁTICOS4.docx"

# # Extrai texto e imagens do arquivo DOCX
# texto, sucesso, erro_msg = extract_text_and_images_from_docx(docx_path)

# if sucesso:
#     # print("Texto completo extraído com sucesso:")
#     # print(texto)
#     with open(r"D:\Repositorios\ocr-api\Versao7\arquivosGerais\docx\output.txt", "w", encoding="utf-8") as file:
#         file.write(texto)
# else:
#     print(f"Erro durante a extração: {erro_msg}")



# #Caminho ou conteúdo do arquivo ODT
# file_path = r"D:\Repositorios\ocr-api\Versao7\csvs\39_8_13022020130620_representacaoinicialop019.2020 (1).odt"

# # Extrai texto do arquivo
# texto, sucesso, erro_msg = extract_text_from_odt(file_path)

# if sucesso:
#     print("Texto extraído com sucesso:")
#     print(texto)
# else:
#     print(f"Erro durante a extração: {erro_msg}")


# import tempfile

# temp_dir = tempfile.gettempdir()
# print(f"Diretório temporário para o usuário atual: {temp_dir}")





# # Caminho para o arquivo PDF
# pdf_path = r"D:\Repositorios\ocr-api\Versao7\arquivosGerais\OP_39_2023_IP__IP 3838-2022.pdf"
# output_path = r"D:\Repositorios\ocr-api\Versao7\arquivosGerais"

# # Extrai texto e imagens do arquivo DOCX
# texto, sucesso, erro_msg = extract_text_from_pdf_content(pdf_path)

# if sucesso:
#     with open(output_path + "output.txt", "w", encoding="latin1") as file:
#         file.write(texto)
#     # with open("output.txt", "w", encoding="utf-8") as file:
#     #     file.write(texto)
#     print("Texto completo extraído com sucesso:")
#     #print(texto)
# else:
#     print(f"Erro durante a extração: {erro_msg}")


# ods_path = r"D:\Repositorios\ocr-api\Versao7\csvs\OP_96_20222_OP_SENHA_VG__TIM2.ods"

# ods_extractor = ODSExtractor(ods_path)
# extracted_text, success, error = ods_extractor.process()

# if success:
#     print(extracted_text)
# else:
#     print(f"Erro: {error}")



# from docling.document_converter import DocumentConverter

# source = "caminho_ou_URL_do_documento"
# converter = DocumentConverter()
# result = converter.convert(source)
# print(result.document.export_to_markdown())


def extract_text_from_xlsx(file_path_or_content):
    """
    Extrai o texto de todas as planilhas de um arquivo .xlsx.

    Parâmetros:
    - file_path_or_content (str ou bytes): Caminho para o arquivo .xlsx ou conteúdo em bytes.

    Retorna:
    - texto_extraido (str): O texto extraído de todas as planilhas.

    Lança:
    - FileNotFoundError: Se o arquivo não existir.
    - ValueError: Se o arquivo não for um arquivo Excel válido.
    - TypeError: Se o parâmetro 'file_path_or_content' não for do tipo esperado.
    - Exception: Para outros erros inesperados.
    """
    if isinstance(file_path_or_content, bytes):
        # Conteúdo em bytes
        file_content = BytesIO(file_path_or_content)
        excel_file = file_content
    elif isinstance(file_path_or_content, str):
        # Caminho do arquivo
        if not os.path.exists(file_path_or_content):
            raise FileNotFoundError(f"O arquivo '{file_path_or_content}' não existe.")
        excel_file = file_path_or_content
    else:
        raise TypeError("O parâmetro 'file_path_or_content' deve ser um caminho de arquivo (str) ou conteúdo em bytes.")

    try:
        # Lê todas as planilhas do arquivo Excel
        df_dict = pd.read_excel(excel_file, sheet_name=None, engine='openpyxl')
        texto_extraido = ''

        for nome_planilha, df in df_dict.items():
            texto_extraido += f"=== Planilha: {nome_planilha} ===\n"
            texto_extraido += df.to_string(index=False)
            texto_extraido += '\n\n'

        return texto_extraido

    except FileNotFoundError:
        raise FileNotFoundError(f"O arquivo '{file_path_or_content}' não foi encontrado.")
    except ValueError as ve:
        raise ValueError(f"Erro ao ler o arquivo xlsx: {ve}")
    except KeyError as ke:
        if "xl/sharedStrings.xml" in str(ke):
            # Tratamento específico para a ausência de sharedStrings.xml
            resultado = extrair_texto_xlsx(file_path_or_content)
            return resultado
        raise Exception(f"Erro inesperado ao processar o arquivo: {ke}")
    except Exception as e:
        raise Exception(f"Erro inesperado ao processar o arquivo: {e}")


#resultado = extract_text_from_xlsx(file_path)
#print(resultado)


# import zipfile
# import xml.etree.ElementTree as ET


# def extrair_texto_xlsx(caminho_arquivo):
#     """
#     Extrai texto de um arquivo .xlsx, lidando com a ausência de xl/sharedStrings.xml.
#     A função explora arquivos sheetN.xml na pasta xl/worksheets.

#     Args:
#         caminho_arquivo (str): Caminho para o arquivo .xlsx.

#     Returns:
#         dict: Um dicionário com o texto extraído de cada planilha.
#     """
#     def extrair_texto(elemento):
#         """
#         Extrai texto de um elemento XML, incluindo filhos e tails.
#         """
#         texto = elemento.text or ''
#         for filho in elemento:
#             texto += extrair_texto(filho)
#             if filho.tail:
#                 texto += filho.tail
#         return texto

#     resultados = {}

#     try:
#         with zipfile.ZipFile(caminho_arquivo, 'r') as arquivo_zip:
#             # Lista arquivos no ZIP
#             arquivos = arquivo_zip.namelist()
#             sheets = [arq for arq in arquivos if arq.startswith('xl/worksheets/sheet') and arq.endswith('.xml')]

#             if not sheets:
#                 return {"Erro": "Nenhuma planilha encontrada no arquivo .xlsx"}

#             for sheet in sheets:
#                 try:
#                     with arquivo_zip.open(sheet) as arquivo_xml:
#                         arvore = ET.parse(arquivo_xml)
#                         raiz = arvore.getroot()
#                         texto_extraido = extrair_texto(raiz)
#                         resultados[sheet] = texto_extraido.strip() or "Sem texto extraído"
#                 except Exception as e:
#                     resultados[sheet] = f"Erro ao processar: {e}"

#     except FileNotFoundError:
#         resultados["Erro"] = f"Arquivo não encontrado: {caminho_arquivo}"
#     except zipfile.BadZipFile:
#         resultados["Erro"] = "O arquivo fornecido não é um arquivo ZIP válido."
#     except Exception as e:
#         resultados["Erro"] = f"Erro inesperado: {e}"

#     return resultados


# # Exemplo de uso:
# caminho_arquivo = file_path
# resultado = extrair_texto_xlsx(caminho_arquivo)

# # Imprime os resultados
# for planilha, conteudo in resultado.items():
#     print(f"Planilha: {planilha}")
#     print(f"Conteúdo:\n{conteudo}\n{'-'*80}")


import zipfile
import xml.etree.ElementTree as ET

def extrair_texto_xlsx(caminho_arquivo):
    """
    Extrai texto de um arquivo .xlsx, lidando com a ausência de xl/sharedStrings.xml.
    A função explora arquivos sheetN.xml na pasta xl/worksheets e insere espaços entre os textos das colunas.

    Args:
        caminho_arquivo (str): Caminho para o arquivo .xlsx.

    Returns:
        dict: Um dicionário com o texto extraído de cada planilha.
    """
    def extrair_texto_com_espacos(elemento):
        """
        Extrai texto de um elemento XML com espaços entre colunas.
        """
        texto = []
        for filho in elemento:
            if filho.text:
                texto.append(filho.text.strip())
            if filho.tail:
                texto.append(filho.tail.strip())
            texto.extend(extrair_texto_com_espacos(filho))
        return texto

    resultados = {}

    try:
        with zipfile.ZipFile(caminho_arquivo, 'r') as arquivo_zip:
            # Lista arquivos no ZIP
            arquivos = arquivo_zip.namelist()
            sheets = [arq for arq in arquivos if arq.startswith('xl/worksheets/sheet') and arq.endswith('.xml')]

            if not sheets:
                return {"Erro": "Nenhuma planilha encontrada no arquivo .xlsx"}

            for sheet in sheets:
                try:
                    with arquivo_zip.open(sheet) as arquivo_xml:
                        arvore = ET.parse(arquivo_xml)
                        raiz = arvore.getroot()
                        
                        # Extrai texto com espaços entre as colunas
                        texto_extraido = " ".join(extrair_texto_com_espacos(raiz)).strip()
                        resultados[sheet] = texto_extraido or "Sem texto extraído"
                except Exception as e:
                    resultados[sheet] = f"Erro ao processar: {e}"

    except FileNotFoundError:
        resultados["Erro"] = f"Arquivo não encontrado: {caminho_arquivo}"
    except zipfile.BadZipFile:
        resultados["Erro"] = "O arquivo fornecido não é um arquivo ZIP válido."
    except Exception as e:
        resultados["Erro"] = f"Erro inesperado: {e}"

    return resultados


# Exemplo de uso:
caminho_arquivo = file_path
# resultado = extrair_texto_xlsx(caminho_arquivo)
resultado = extract_text_from_xlsx(caminho_arquivo)
print(resultado)


# # Imprime os resultados
# for planilha, conteudo in resultado.items():
#     print(f"Planilha: {planilha}")
#     print(f"Conteúdo:\n{conteudo}\n{'-'*80}")
