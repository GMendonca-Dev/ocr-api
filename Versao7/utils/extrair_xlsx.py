import zipfile
import xml.etree.ElementTree as ET
import pandas as pd
from io import BytesIO
import os


def extract_text_from_xlsx(file_path_or_content):
    """
    Extrai o texto de todas as planilhas de um arquivo .xlsx.

    Caso o arquivo não possua "xl/sharedStrings.xml", utiliza a função "extrair_texto_xlsx".

    Parâmetros:
    - file_path_or_content (str ou bytes): Caminho para o arquivo .xlsx ou conteúdo em bytes.

    Retorna:
    - texto_extraido (str): O texto extraído de todas as planilhas.

    Lança:
    - FileNotFoundError: Se o arquivo não existir.
    - ValueError: Se o arquivo não for um arquivo Excel válido.
    - TypeError: Se o parâmetro 'file_path_or_content' não for do tipo esperado.
    - Exception: Para outros erros inesperados.
    """
    if isinstance(file_path_or_content, bytes):
        # Conteúdo em bytes
        file_content = BytesIO(file_path_or_content)
        excel_file = file_content
    elif isinstance(file_path_or_content, str):
        # Caminho do arquivo
        if not os.path.exists(file_path_or_content):
            raise FileNotFoundError(f"O arquivo '{file_path_or_content}' não existe.")
        excel_file = file_path_or_content
    else:
        raise TypeError("O parâmetro 'file_path_or_content' deve ser um caminho de arquivo (str) ou conteúdo em bytes.")

    try:
        # Tenta ler todas as planilhas do arquivo Excel
        df_dict = pd.read_excel(excel_file, sheet_name=None, engine='openpyxl')
        texto_extraido = ''

        for nome_planilha, df in df_dict.items():
            texto_extraido += f"=== Planilha: {nome_planilha} ===\n"
            texto_extraido += df.to_string(index=False)
            texto_extraido += '\n\n'

        return texto_extraido

    except KeyError as ke:
        if "xl/sharedStrings.xml" in str(ke):
            # Tratamento específico para a ausência de sharedStrings.xml
            return extrair_texto_xlsx(file_path_or_content)
        raise Exception(f"Erro inesperado ao processar o arquivo: {ke}")
    except Exception as e:
        raise Exception(f"Erro inesperado ao processar o arquivo: {e}")


def extrair_texto_xlsx(caminho_arquivo):
    """
    Extrai texto de um arquivo .xlsx, lidando com a ausência de xl/sharedStrings.xml.
    A função explora arquivos sheetN.xml na pasta xl/worksheets e insere espaços entre os textos das colunas.

    Args:
        caminho_arquivo (str): Caminho para o arquivo .xlsx.

    Returns:
        str: Texto extraído de todas as planilhas, concatenado.
    """
    def extrair_texto_com_espacos(elemento):
        """
        Extrai texto de um elemento XML com espaços entre colunas.
        """
        texto = []
        for filho in elemento:
            if filho.text:
                texto.append(filho.text.strip())
            if filho.tail:
                texto.append(filho.tail.strip())
            texto.extend(extrair_texto_com_espacos(filho))
        return texto

    resultados = []

    try:
        with zipfile.ZipFile(caminho_arquivo, 'r') as arquivo_zip:
            # Lista arquivos no ZIP
            arquivos = arquivo_zip.namelist()
            sheets = [arq for arq in arquivos if arq.startswith('xl/worksheets/sheet') and arq.endswith('.xml')]

            if not sheets:
                return "Nenhuma planilha encontrada no arquivo .xlsx"

            for sheet in sheets:
                try:
                    with arquivo_zip.open(sheet) as arquivo_xml:
                        arvore = ET.parse(arquivo_xml)
                        raiz = arvore.getroot()

                        # Extrai texto com espaços entre as colunas
                        texto_extraido = " ".join(extrair_texto_com_espacos(raiz)).strip()
                        if texto_extraido:
                            resultados.append(texto_extraido)
                except Exception as e:
                    resultados.append(f"Erro ao processar {sheet}: {e}")

    except FileNotFoundError:
        return f"Arquivo não encontrado: {caminho_arquivo}"
    except zipfile.BadZipFile:
        return "O arquivo fornecido não é um arquivo ZIP válido."
    except Exception as e:
        return f"Erro inesperado: {e}"

    return "\n\n".join(resultados)


# file_path = r"D:\Repositorios\ocr-api\Versao7\arquivosGerais\csvs\OP_213_2023_OP_SENHA_VG.xlsx"
# print(extract_text_from_xlsx(file_path))



import os
import pandas as pd
import zipfile
import xml.etree.ElementTree as ET
from io import BytesIO
import xlrd  # Biblioteca para ler arquivos XLS binários
import xlrd
from openpyxl import Workbook



def extract_text_from_xls(xls_file_path):
    """
    Extrai texto de um arquivo .xls, tentando primeiro converter para .xlsx e, se falhar, tenta ler como CSV ou HTML.

    Parâmetros:
    - xls_file_path (str): Caminho para o arquivo .xls.

    Retorna:
    - str: Texto extraído do arquivo.
    """
    try:
        # Tenta abrir o arquivo .xls
        workbook_xls = xlrd.open_workbook(xls_file_path)
        xlsx_file_path = xls_file_path.replace('.xls', '.xlsx')
        workbook_xlsx = Workbook()

        for sheet_index in range(workbook_xls.nsheets):
            sheet_xls = workbook_xls.sheet_by_index(sheet_index)
            sheet_name = sheet_xls.name

            # Cria uma aba correspondente no .xlsx
            sheet_xlsx = workbook_xlsx.create_sheet(title=sheet_name) if sheet_index > 0 else workbook_xlsx.active
            sheet_xlsx.title = sheet_name

            # Copia os dados da aba
            for row in range(sheet_xls.nrows):
                for col in range(sheet_xls.ncols):
                    sheet_xlsx.cell(row=row + 1, column=col + 1, value=sheet_xls.cell_value(row, col))

        # Remove a aba padrão criada automaticamente
        if "Sheet" in workbook_xlsx.sheetnames:
            del workbook_xlsx["Sheet"]

        # Salva o arquivo convertido
        workbook_xlsx.save(xlsx_file_path)

        # Agora, extrai o texto do arquivo .xlsx convertido
        return extract_text_from_xlsx(xlsx_file_path)

    except (ValueError, xlrd.XLRDError) as e:
        print(f"Erro ao abrir o arquivo .xls: {e}. Tentando ler como CSV ou HTML.")  # Log
        resultados = read_file_as_csv_or_html(xls_file_path)
        return resultados

    # except Exception as e:
    #     print(f"Erro ao converter para .xlsx: {e}. Tentando ler como CSV ou HTML.")  # Log
    #     return read_file_as_csv_or_html(xls_file_path)

def read_file_as_csv_or_html(file_path):
    """
    Tenta ler o arquivo como CSV ou HTML e extrair texto.

    Parâmetros:
    - file_path (str): Caminho para o arquivo.

    Retorna:
    - str: Texto extraído do arquivo.
    """
    try:
        # Tenta ler como CSV
        df = pd.read_csv(file_path)
        return df.to_string(index=False)

    except Exception as e:
        print(f"Erro ao ler como CSV: {e}. Tentando ler como HTML.")  # Log
        try:
            # Tenta ler como HTML
            df = pd.read_html(file_path)
            return "\n".join([df.to_string(index=False) for df in df])

        except Exception as e:
            print(f"Erro ao ler como HTML: {e}")  # Log
            return "Não foi possível extrair texto do arquivo"


file_path = r"D:\Repositorios\ocr-api\Versao7\arquivosGerais\csvs\OP_266_2019_ANOT_AN_ARQUIVOS DIVERSOS INICIAIS8.xls"
#print(extract_text_from_xls(file_path))


try:
    texto_extraido = extract_text_from_xls(file_path)
    print(texto_extraido)
except Exception as e:
    print(f"Erro ao extrair texto: {e}")